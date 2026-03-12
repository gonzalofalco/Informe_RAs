#!/usr/bin/env python3
"""Automatiza el login en usuarios.telecentro.net.ar usando Playwright.

Uso:
  export TELECENTRO_USER='gfalco'
  export TELECENTRO_PASS='...'
    export CRM_USER='gfalco'
    export CRM_PASS='...'
  python3 login_telecentro.py
"""

import os
import json
import re
import base64
import csv
import shutil
import socket
import sys
import ipaddress
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import Playwright, TimeoutError as PlaywrightTimeoutError, sync_playwright

LOGIN_URL = "https://usuarios.telecentro.net.ar/logIn.php"
CRM_URL_DEFAULT = "http://crm.telecentro.local//MembersLogin.aspx"
CRM_RECLAMO_URL_DEFAULT = "http://crm.telecentro.local//Cliente/ReclamoAdministrativo/ReclamoCierre.aspx?SubMenu=475"


def get_timeout_ms(env_name: str, default_value: int) -> int:
    raw = os.getenv(env_name, "").strip()
    if not raw:
        return default_value
    try:
        value = int(raw)
        return value if value > 0 else default_value
    except ValueError:
        return default_value


def get_lookback_days(default_value: int = 7) -> int:
    raw = os.getenv("TELECENTRO_LOOKBACK_DAYS", "").strip()
    if not raw:
        return default_value
    try:
        value = int(raw)
        return value if value >= 0 else default_value
    except ValueError:
        return default_value


UI_TIMEOUT_MS = get_timeout_ms("TELECENTRO_UI_TIMEOUT_MS", 15000)
RECLAMO_WAIT_MS = get_timeout_ms("TELECENTRO_RECLAMO_WAIT_MS", 45000)
DOWNLOAD_TIMEOUT_MS = get_timeout_ms("TELECENTRO_DOWNLOAD_TIMEOUT_MS", 120000)


def get_target_classifications() -> list[str]:
    """Devuelve clasificaciones objetivo desde env, con default compatible."""
    raw = os.getenv("TELECENTRO_CLASIFICACIONES", "").strip()
    if not raw:
        return ["CNOC - BANDA ANCHA"]

    # Permite separar por | o por coma para facilitar configuración en .env
    parts = [item.strip() for item in re.split(r"[|,]", raw) if item.strip()]
    return parts or ["CNOC - BANDA ANCHA"]


def classification_filename(text: str) -> str:
    """Convierte una clasificación en nombre de archivo legible y seguro."""
    base = re.sub(r"[\\/:*?\"<>|]", " ", (text or "").strip())
    base = re.sub(r"\s+", " ", base).strip()
    return base or "clasificacion"


def normalize_text(value: str) -> str:
    """Normaliza texto para comparaciones tolerantes (acentos/case/simbolos)."""
    text = (value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_filtered_csv_from_report(report_path: Path, classification: str, download_dir: Path) -> Path:
    """Genera CSV filtrado por palabras clave para una clasificacion."""
    try:
        from openpyxl import load_workbook
    except ImportError as err:
        raise RuntimeError(
            "Falta dependencia 'openpyxl' para generar CSV. "
            "Ejecuta: pip install -r requirements.txt"
        ) from err

    wb = load_workbook(filename=str(report_path), read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell).strip() for cell in row])

    if not rows:
        raise RuntimeError(f"El reporte esta vacio: {report_path}")

    required_headers = {
        "reclamo": "RECLAMO ID",
        "cliente": "CLIENTE NUMERO",
        "motivo": "ULTIMO COMENTARIO",
    }

    header_idx = -1
    col_reclamo = -1
    col_cliente = -1
    col_motivo = -1

    for ridx, row in enumerate(rows):
        normalized_cells = [normalize_text(cell) for cell in row]

        def _find_col(key: str) -> int:
            target = required_headers[key]

            # 1) Match exacto primero.
            for cidx, cell in enumerate(normalized_cells):
                if cell == target:
                    return cidx

            # 2) Fallback por inclusion con reglas para evitar falsos positivos.
            for cidx, cell in enumerate(normalized_cells):
                if target not in cell:
                    continue
                if key == "motivo" and "USUARIO" in cell:
                    # Evita mapear a "USUARIO ULTIMO COMENTARIO".
                    continue
                return cidx

            return -1

        tmp_reclamo = _find_col("reclamo")
        tmp_cliente = _find_col("cliente")
        tmp_motivo = _find_col("motivo")

        if tmp_reclamo >= 0 and tmp_cliente >= 0 and tmp_motivo >= 0:
            header_idx = ridx
            col_reclamo = tmp_reclamo
            col_cliente = tmp_cliente
            col_motivo = tmp_motivo
            break

    if header_idx < 0:
        raise RuntimeError(
            "No se encontro encabezado con columnas requeridas: "
            "RECLAMO ID, CLIENTE NUMERO, ULTIMO COMENTARIO"
        )

    keywords = [
        "FALLA MASIVA",
        "PROBLEMA RESUELTO",
        "ENVIAR VT",
        "SIN FALLA",
    ]

    def _extract_keyword_label(comment_text: str) -> str:
        norm = normalize_text(comment_text)
        for keyword in keywords:
            if keyword in norm:
                return keyword
        return ""

    out_rows = []
    for row in rows[header_idx + 1 :]:
        max_col = max(col_reclamo, col_cliente, col_motivo)
        if len(row) <= max_col:
            continue

        reclamo_id = row[col_reclamo].strip()
        cliente_numero = row[col_cliente].strip()
        ultimo_comentario = row[col_motivo].strip()

        if not (reclamo_id or cliente_numero or ultimo_comentario):
            continue

        motivo_keyword = _extract_keyword_label(ultimo_comentario)
        if not motivo_keyword:
            continue

        out_rows.append([reclamo_id, cliente_numero, motivo_keyword])

    csv_path = download_dir / f"{classification_filename(classification)} - filtrado.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["RECLAMO ID", "CLIENTE NUMERO", "Motivo"])
        writer.writerows(out_rows)

    print(
        "CSV filtrado generado para "
        f"'{classification}': {csv_path} (filas={len(out_rows)})"
    )
    return csv_path


def extract_original_csv_from_report(report_path: Path, classification: str, download_dir: Path) -> Path:
    """Convierte el reporte XLSX completo a CSV original (sin filtrado)."""
    try:
        from openpyxl import load_workbook
    except ImportError as err:
        raise RuntimeError(
            "Falta dependencia 'openpyxl' para generar CSV. "
            "Ejecuta: pip install -r requirements.txt"
        ) from err

    wb = load_workbook(filename=str(report_path), read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell).strip() for cell in row])

    if not rows:
        raise RuntimeError(f"El reporte esta vacio: {report_path}")

    # Busca encabezado real para evitar metadata repetida de la cabecera visual del reporte.
    header_idx = -1
    for ridx, row in enumerate(rows):
        normalized = [normalize_text(cell) for cell in row]
        if any(cell == "RECLAMO ID" for cell in normalized):
            header_idx = ridx
            break

    data_rows = rows[header_idx:] if header_idx >= 0 else rows
    csv_path = download_dir / f"{classification_filename(classification)} - original.csv"

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(data_rows)

    print(
        "CSV original generado para "
        f"'{classification}': {csv_path} (filas={max(len(data_rows) - 1, 0)})"
    )
    return csv_path


def iter_targets(page):
    """Devuelve página principal + iframes para buscar controles dinámicos."""
    return [page] + [frame for frame in page.frames if frame != page.main_frame]


def dump_debug_artifacts(page, prefix: str) -> None:
    artifacts_dir = Path("artifacts")
    artifacts_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    html_path = artifacts_dir / f"{prefix}_{ts}.html"
    png_path = artifacts_dir / f"{prefix}_{ts}.png"

    html_path.write_text(page.content(), encoding="utf-8")
    page.screenshot(path=str(png_path), full_page=True)
    print(f"Debug guardado: {html_path}")
    print(f"Debug guardado: {png_path}")

    # Guarda también HTML de cada iframe para diagnóstico fino.
    for idx, frame in enumerate(page.frames):
        try:
            frame_html = artifacts_dir / f"{prefix}_{ts}_frame_{idx}.html"
            frame_html.write_text(frame.content(), encoding="utf-8")
            print(f"Debug frame guardado: {frame_html} ({frame.url})")
        except PlaywrightError:
            continue


def get_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        print(f"Falta variable de entorno obligatoria: {name}")
        sys.exit(1)
    return value


def fill_first_visible(page, selectors, value: str, label: str) -> None:
    for target in iter_targets(page):
        for selector in selectors:
            locator = target.locator(selector)
            if locator.count() == 0:
                continue
            try:
                locator.first.wait_for(state="visible", timeout=UI_TIMEOUT_MS)
                locator.first.fill(value)
                print(f"Campo {label} completado con selector: {selector} en {target.url}")
                return
            except PlaywrightTimeoutError:
                continue

    raise RuntimeError(f"No se encontró un campo visible para: {label}")


def _build_targets_from_roots(roots):
    """Expande páginas/frame roots en una sola lista de targets navegables."""
    targets = []
    for root in roots:
        # Page y Frame exponen .locator(). Para Page agregamos también sus iframes.
        targets.append(root)
        frames = getattr(root, "frames", None)
        main_frame = getattr(root, "main_frame", None)
        if isinstance(frames, list) and main_frame is not None:
            for frame in frames:
                if frame != main_frame:
                    targets.append(frame)
    return targets


def click_first_visible(page, selectors, label: str, roots=None) -> None:
    targets = _build_targets_from_roots(roots) if roots else iter_targets(page)
    for target in targets:
        for selector in selectors:
            locator = target.locator(selector)
            if locator.count() == 0:
                continue
            try:
                locator.first.wait_for(state="visible", timeout=UI_TIMEOUT_MS)
                locator.first.click()
                print(f"Click en {label} con selector: {selector} en {target.url}")
                return
            except PlaywrightTimeoutError:
                continue

    raise RuntimeError(f"No se encontró un botón visible para: {label}")


def click_first_visible_retry(page, selectors, label: str, timeout_ms: int = 30000, roots=None) -> bool:
    """Variante no-exception: reintenta click hasta timeout y devuelve True/False."""
    started = datetime.now()
    while (datetime.now() - started).total_seconds() * 1000 < timeout_ms:
        targets = _build_targets_from_roots(roots) if roots else iter_targets(page)
        for target in targets:
            for selector in selectors:
                locator = target.locator(selector)
                if locator.count() == 0:
                    continue
                try:
                    locator.first.wait_for(state="visible", timeout=1200)
                    locator.first.click()
                    print(f"Click en {label} con selector: {selector} en {target.url}")
                    return True
                except PlaywrightError:
                    continue
        page.wait_for_timeout(400)
    return False


def select_first_visible(page, selectors, option_label: str, label: str) -> None:
    for target in iter_targets(page):
        for selector in selectors:
            locator = target.locator(selector)
            if locator.count() == 0:
                continue

            total = locator.count()
            for idx in range(total):
                item = locator.nth(idx)
                try:
                    # Permite selects ocultos que sí están vinculados al form.
                    item.wait_for(state="attached", timeout=UI_TIMEOUT_MS)
                    item.select_option(label=option_label)
                    print(f"Selección {label} aplicada con selector: {selector} -> {option_label} en {target.url}")
                    return
                except (PlaywrightTimeoutError, PlaywrightError):
                    continue

    # Fallback por valor textual exacto de option.
    for target in iter_targets(page):
        for selector in selectors:
            locator = target.locator(selector)
            if locator.count() == 0:
                continue
            try:
                locator.first.select_option(option_label)
                print(f"Selección {label} aplicada por valor: {option_label} en {target.url}")
                return
            except PlaywrightError:
                continue

    raise RuntimeError(f"No se encontró un selector visible para: {label}")


def select_option_in_any_select(page, option_label: str, label: str) -> None:
    """Fallback: busca cualquier <select> que contenga la opción y la selecciona."""
    for target in iter_targets(page):
        selects = target.locator("select")
        count = selects.count()
        for idx in range(count):
            item = selects.nth(idx)
            try:
                item.wait_for(state="attached", timeout=UI_TIMEOUT_MS)
                item.select_option(label=option_label)
                print(f"Selección {label} aplicada en select genérico ({target.url}) -> {option_label}")
                return
            except PlaywrightError:
                continue

    raise RuntimeError(f"No se encontró ninguna opción '{option_label}' para {label}")


def select_estado_cerrado(page) -> None:
    """Selecciona Estado=Cerrado en el filtro correcto y valida el valor final."""
    selectors = [
        # Prioriza el select asociado visualmente a la etiqueta "Estado".
        "xpath=//*[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'estado')]/following::select[1]",
        "select#ddlEstado",
        "select[id*='Estado']",
        "select[name*='Estado']",
    ]

    for target in iter_targets(page):
        for selector in selectors:
            try:
                locator = target.locator(selector)
            except PlaywrightError:
                continue
            if locator.count() == 0:
                continue

            for idx in range(locator.count()):
                item = locator.nth(idx)
                try:
                    item.wait_for(state="attached", timeout=UI_TIMEOUT_MS)
                    item.select_option(label="Cerrado")
                    # Fuerza onchange para UIs que dependen de eventos JS.
                    item.evaluate(
                        """
                        (el) => {
                            el.dispatchEvent(new Event('input', { bubbles: true }));
                            el.dispatchEvent(new Event('change', { bubbles: true }));
                            el.dispatchEvent(new Event('blur', { bubbles: true }));
                        }
                        """
                    )
                    selected = item.evaluate(
                        """
                        (el) => {
                            const opt = el.options && el.options[el.selectedIndex];
                            return (opt && opt.textContent ? opt.textContent : '').trim();
                        }
                        """
                    )
                    if "cerrado" in (selected or "").strip().lower():
                        print(f"Estado validado en 'Cerrado' con selector: {selector} en {target.url}")
                        return
                except PlaywrightError:
                    continue

    raise RuntimeError("No se pudo dejar el filtro Estado en 'Cerrado'.")


def wait_results_refresh_after_search(page) -> None:
    """Espera refresh de resultados tras 'Buscar por Estado' para evitar exportar datos viejos."""
    for target in iter_targets(page):
        try:
            overlay = target.locator("#loadingOverlay")
            if overlay.count() == 0:
                continue
            try:
                overlay.first.wait_for(state="visible", timeout=2500)
            except PlaywrightTimeoutError:
                pass
            try:
                overlay.first.wait_for(state="hidden", timeout=25000)
            except PlaywrightTimeoutError:
                pass
        except PlaywrightError:
            continue

    # Buffer extra para asegurar que tabla/filtros terminaron de sincronizar.
    page.wait_for_timeout(2000)


def fill_date_input_by_order(page, index: int, value: str, label: str) -> None:
    """Fallback para fechas cuando no hay ids estables (toma N-esimo campo de fecha)."""
    candidates_selectors = [
        "input[id*='Fecha' i]",
        "input[id*='Fec' i]",
        "input[name*='Fecha' i]",
        "input[name*='Fec' i]",
        "input[placeholder*='Fecha' i]",
        "input[type='date']",
    ]

    for target in iter_targets(page):
        seen = []
        for selector in candidates_selectors:
            try:
                loc = target.locator(selector)
                for i in range(loc.count()):
                    item = loc.nth(i)
                    try:
                        item.wait_for(state="visible", timeout=1200)
                        seen.append(item)
                    except PlaywrightTimeoutError:
                        continue
            except PlaywrightError:
                continue

        if len(seen) > index:
            item = seen[index]
            try:
                item.fill(value)
                print(f"Campo {label} completado por orden ({index}) en {target.url}")
                return
            except PlaywrightError:
                # Algunas variantes de CRM exponen un input deshabilitado; forzamos seteo y eventos.
                try:
                    item.evaluate(
                        """
                        (el, v) => {
                            try { el.removeAttribute('disabled'); } catch (e) {}
                            el.value = v;
                            el.dispatchEvent(new Event('input', { bubbles: true }));
                            el.dispatchEvent(new Event('change', { bubbles: true }));
                            el.dispatchEvent(new Event('blur', { bubbles: true }));
                        }
                        """,
                        value,
                    )
                    print(f"Campo {label} completado por JS fallback ({index}) en {target.url}")
                    return
                except PlaywrightError:
                    pass

    raise RuntimeError(f"No se pudo completar {label} por fallback de orden")


def get_first_visible_input_value(page, selectors) -> str:
    """Lee el valor del primer input visible que matchee los selectores."""
    for target in iter_targets(page):
        for selector in selectors:
            try:
                locator = target.locator(selector)
            except PlaywrightError:
                continue
            if locator.count() == 0:
                continue
            for idx in range(locator.count()):
                item = locator.nth(idx)
                try:
                    item.wait_for(state="visible", timeout=800)
                    return (item.input_value() or "").strip()
                except PlaywrightError:
                    continue
    return ""


def get_first_visible_select_text(page, selectors) -> str:
    """Lee el texto seleccionado del primer select visible."""
    for target in iter_targets(page):
        for selector in selectors:
            try:
                locator = target.locator(selector)
            except PlaywrightError:
                continue
            if locator.count() == 0:
                continue
            for idx in range(locator.count()):
                item = locator.nth(idx)
                try:
                    item.wait_for(state="visible", timeout=800)
                    text = item.evaluate(
                        """
                        (el) => {
                            const opt = el.options && el.options[el.selectedIndex];
                            return (opt && opt.textContent ? opt.textContent : '').trim();
                        }
                        """
                    )
                    if text:
                        return text
                except PlaywrightError:
                    continue
    return ""


def force_set_date_filters(page, fecha_desde: str, fecha_hasta: str) -> None:
    """Fuerza seteo de fechas por JS sobre campos Desde/Hasta cuando la UI no refleja fill()."""
    script = """
    (vals) => {
        const all = Array.from(document.querySelectorAll('input'));
        const norm = (s) => (s || '').toLowerCase();
        const isFrom = (el) => {
            const k = norm(el.id) + ' ' + norm(el.name) + ' ' + norm(el.placeholder);
            return k.includes('fechadesde') || k.includes('fecdesde') || k.includes('desde');
        };
        const isTo = (el) => {
            const k = norm(el.id) + ' ' + norm(el.name) + ' ' + norm(el.placeholder);
            return k.includes('fechahasta') || k.includes('fechasta') || k.includes('fechah') || k.includes('fechahst') || k.includes('fecha_hasta') || k.includes('fechahasta') || k.includes('fechto') || k.includes('fechafin') || k.includes('hasta');
        };
        const setVal = (el, v) => {
            try { el.removeAttribute('readonly'); } catch (e) {}
            try { el.removeAttribute('disabled'); } catch (e) {}
            el.value = v;
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
            el.dispatchEvent(new Event('blur', { bubbles: true }));
        };
        let touchedFrom = 0;
        let touchedTo = 0;
        for (const el of all) {
            if (isFrom(el)) { setVal(el, vals.desde); touchedFrom++; }
            if (isTo(el)) { setVal(el, vals.hasta); touchedTo++; }
        }
        return { touchedFrom, touchedTo };
    }
    """
    for target in iter_targets(page):
        try:
            result = target.evaluate(script, {"desde": fecha_desde, "hasta": fecha_hasta})
            if result:
                print(
                    "Forzado JS de fechas en "
                    f"{target.url} (desde={result.get('touchedFrom', 0)}, hasta={result.get('touchedTo', 0)})"
                )
        except PlaywrightError:
            continue


def ensure_filters_consistency(page, expected_classification: str, fecha_desde: str, fecha_hasta: str) -> None:
    """Valida que clasificación/estado/fechas queden aplicados antes de buscar."""
    class_text = get_first_visible_select_text(
        page,
        [
            "select#ddlClasificacion",
            "select[id*='Clasificacion']",
            "select[name*='Clasificacion']",
        ],
    )
    estado_text = get_first_visible_select_text(
        page,
        [
            "xpath=//*[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'estado')]/following::select[1]",
            "select#ddlEstado",
            "select[id*='Estado']",
            "select[name*='Estado']",
        ],
    )
    desde_value = get_first_visible_input_value(
        page,
        [
            "input#txtFechaDesde",
            "input#txtFecDesde",
            "input[id*='FechaDesde']",
            "input[id*='FecDesde']",
            "input[name*='FechaDesde']",
            "input[name*='FecDesde']",
            "input[placeholder*='Desde' i]",
        ],
    )
    hasta_value = get_first_visible_input_value(
        page,
        [
            "input#txtFechaHasta",
            "input#txtFecHasta",
            "input[id*='FechaHasta']",
            "input[id*='FecHasta']",
            "input[name*='FechaHasta']",
            "input[name*='FecHasta']",
            "input[placeholder*='Hasta' i]",
        ],
    )

    ok_class = expected_classification.lower() in (class_text or "").strip().lower()
    ok_estado = "cerrado" in (estado_text or "").strip().lower()
    ok_desde = (desde_value or "").strip() == fecha_desde
    ok_hasta = (hasta_value or "").strip() == fecha_hasta

    if ok_class and ok_estado and ok_desde and ok_hasta:
        print(
            "Filtros validados: "
            f"clasificacion='{class_text}', estado='{estado_text}', desde='{desde_value}', hasta='{hasta_value}'"
        )
        return

    # Reintento correctivo de fechas cuando la UI pisa valores tras onchange.
    if not (ok_desde and ok_hasta):
        force_set_date_filters(page, fecha_desde, fecha_hasta)
        desde_value = get_first_visible_input_value(
            page,
            [
                "input#txtFechaDesde",
                "input#txtFecDesde",
                "input[id*='FechaDesde']",
                "input[id*='FecDesde']",
                "input[name*='FechaDesde']",
                "input[name*='FecDesde']",
                "input[placeholder*='Desde' i]",
            ],
        )
        hasta_value = get_first_visible_input_value(
            page,
            [
                "input#txtFechaHasta",
                "input#txtFecHasta",
                "input[id*='FechaHasta']",
                "input[id*='FecHasta']",
                "input[name*='FechaHasta']",
                "input[name*='FecHasta']",
                "input[placeholder*='Hasta' i]",
            ],
        )
        ok_desde = (desde_value or "").strip() == fecha_desde
        ok_hasta = (hasta_value or "").strip() == fecha_hasta

    if not (ok_class and ok_estado and ok_desde and ok_hasta):
        dump_debug_artifacts(page, "crm_filters_inconsistent")
        raise RuntimeError(
            "Los filtros quedaron inconsistentes antes de buscar: "
            f"clasificacion='{class_text}' esperado='{expected_classification}', "
            f"estado='{estado_text}' esperado='Cerrado', "
            f"desde='{desde_value}' esperado='{fecha_desde}', "
            f"hasta='{hasta_value}' esperado='{fecha_hasta}'"
        )


def iframe_requires_login(page) -> bool:
    """Indica si el iframe mainFrame quedó en pantalla de login de CRMWeb."""
    crm_frame = page.frame(name="mainFrame")
    if not crm_frame:
        return False

    try:
        return (
            crm_frame.locator("#txtUsuario").count() > 0
            and crm_frame.locator("#txtPassword").count() > 0
            and crm_frame.locator("#btnAceptar, input[name='btnAceptar']").count() > 0
        )
    except PlaywrightError:
        return False


def get_iframe_login_error(page) -> str:
    """Devuelve mensaje de error visible en login embebido, si existe."""
    crm_frame = page.frame(name="mainFrame")
    if not crm_frame:
        return ""

    try:
        locator = crm_frame.locator("#lblError")
        if locator.count() == 0:
            return ""
        text = locator.first.inner_text().strip()
        return text
    except PlaywrightError:
        return ""


def login_iframe_if_needed(page, crm_user: str, crm_pass: str) -> None:
    """Si mainFrame muestra login, lo completa para continuar hacia filtros."""
    if not iframe_requires_login(page):
        return

    crm_frame = page.frame(name="mainFrame")
    if not crm_frame:
        return

    for attempt in [1, 2]:
        print(f"Detectado login embebido en mainFrame; intento {attempt}/2.")

        # En CRMWeb suele venir usuario preasignado por sesión de red.
        user_locator = crm_frame.locator("#txtUsuario")
        current_user = ""
        if user_locator.count() > 0:
            try:
                current_user = (user_locator.first.input_value() or "").strip()
            except PlaywrightError:
                current_user = ""

        if not current_user:
            crm_frame.fill("#txtUsuario", crm_user)

        crm_frame.fill("#txtPassword", crm_pass)
        crm_frame.click("#btnAceptar, input[name='btnAceptar']")
        page.wait_for_timeout(3500)

        if not iframe_requires_login(page):
            print("Login embebido completado correctamente.")
            return

        iframe_error = get_iframe_login_error(page)
        if iframe_error:
            print(f"Error en login embebido CRMWeb: {iframe_error}")

        # Reintento: recarga la URL de reclamo para regenerar token/frame state.
        if attempt == 1:
            print("Reintentando flujo de reclamo para regenerar contexto de iframe...")
            page.reload(wait_until="domcontentloaded")
            page.wait_for_timeout(2500)
            crm_frame = page.frame(name="mainFrame")
            if not crm_frame:
                return


def page_requires_login(page) -> bool:
    """Indica si la página actual es login directo de CRM/BILLING."""
    try:
        return (
            page.locator("#txtUsuario").count() > 0
            and page.locator("#txtPassword").count() > 0
            and page.locator("#btnAceptar, input[name='btnAceptar']").count() > 0
        )
    except PlaywrightError:
        return False


def get_page_login_error(page) -> str:
    try:
        locator = page.locator("#lblError")
        if locator.count() == 0:
            return ""
        return locator.first.inner_text().strip()
    except PlaywrightError:
        return ""


def login_page_if_needed(page, crm_user: str, crm_pass: str) -> bool:
    """Realiza login cuando CRM muestra formulario directo en la página principal."""
    if not page_requires_login(page):
        return False

    print("Detectado login directo CRMWeb en página principal; intentando autenticar...")

    try:
        user_input = page.locator("#txtUsuario")
        if user_input.count() > 0:
            # Forzamos usuario objetivo para evitar cuentas precompletadas incorrectas.
            try:
                user_input.first.fill(crm_user)
            except PlaywrightError:
                user_input.first.evaluate(
                    """
                    (el, v) => {
                        try { el.removeAttribute('readonly'); } catch (e) {}
                        try { el.removeAttribute('disabled'); } catch (e) {}
                        el.value = v;
                        el.dispatchEvent(new Event('input', { bubbles: true }));
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                    }
                    """,
                    crm_user,
                )

        page.fill("#txtPassword", crm_pass)
        page.click("#btnAceptar, input[name='btnAceptar']")
    except PlaywrightError as err:
        print(f"No se pudo completar login directo CRMWeb: {err}")
        return False

    try:
        page.wait_for_load_state("networkidle", timeout=UI_TIMEOUT_MS)
    except PlaywrightTimeoutError:
        pass

    err_txt = get_page_login_error(page)
    if err_txt:
        print(f"Error en login directo CRMWeb: {err_txt}")

    return not page_requires_login(page)


def has_parking_redirect(page) -> bool:
    """Detecta si algún frame cayó en dominio externo de parking/forwarding."""
    for frame in page.frames:
        url = (frame.url or "").lower()
        if "instantfwding.com" in url:
            return True
    return False


def crmweb_stuck_on_root(page) -> bool:
    """Detecta when crmweb carga solo home (#/) y no el módulo ReclamoCierre."""
    crmweb_frames = [
        (frame.url or "").lower()
        for frame in page.frames
        if "crmweb.telecentro.local" in (frame.url or "").lower()
    ]
    if not crmweb_frames:
        return False

    has_reclamo_route = any("reclamocierre" in url for url in crmweb_frames)
    has_root_route = any(url.endswith("#/") or url.endswith("/#") for url in crmweb_frames)
    return has_root_route and not has_reclamo_route


def recover_mainframe_reclamo_route(page) -> bool:
    """Si mainFrame quedó en '#/', intenta reabrir la ruta tokenizada de ReclamoCierre."""
    if not crmweb_stuck_on_root(page):
        return False

    src = get_mainframe_src(page)
    if not src or "reclamocierre" not in src.lower():
        print("No hay src tokenizado de ReclamoCierre para recuperar mainFrame.")
        return False

    frame = page.frame(name="mainFrame")
    if not frame:
        print("No se encontró mainFrame para recuperar ruta de ReclamoCierre.")
        return False

    print("mainFrame quedó en '#/'; reintentando ruta tokenizada de ReclamoCierre...")
    try:
        frame.goto(src, wait_until="domcontentloaded")
        page.wait_for_timeout(2000)
    except PlaywrightError as err:
        print(f"Falló recuperación de ruta en mainFrame: {err}")
        return False

    return not crmweb_stuck_on_root(page)


def get_mainframe_src(page) -> str:
    """Obtiene el src actual del iframe principal si existe."""
    selectors = [
        "iframe[name='mainFrame']",
        "iframe#mainFrame",
        "iframe[id*='mainFrame']",
    ]
    for selector in selectors:
        try:
            locator = page.locator(selector)
            if locator.count() == 0:
                continue
            src = locator.first.get_attribute("src") or ""
            src = src.strip()
            if src:
                return src
        except PlaywrightError:
            continue
    return ""


def try_reclamo_fallback_via_token_url(page, crm_internal_host: str) -> bool:
    """Fallback: abre la URL tokenizada del iframe preservando host CRMWeb original."""
    src = get_mainframe_src(page)
    if not src:
        print("No se encontró src de mainFrame para fallback.")
        return False

    # Conserva el host tokenizado original (normalmente crmweb.*).
    # Forzar crm.telecentro.local aquí puede redirigir al backend legacy incorrecto.
    target_url = src

    print(f"Intentando fallback de Reclamo con URL tokenizada: {target_url}")
    try:
        page.goto(target_url, wait_until="domcontentloaded")
    except PlaywrightError as err:
        print(f"Fallback de Reclamo falló al navegar: {err}")
        return False

    try:
        page.wait_for_load_state("networkidle", timeout=UI_TIMEOUT_MS)
    except PlaywrightTimeoutError:
        pass

    if has_parking_redirect(page):
        print("Fallback de Reclamo detectó redirect externo.")
        return False

    return True


def wait_for_reclamo_filters(page) -> None:
    """Espera a que la pantalla de filtros esté disponible antes de interactuar."""
    selectors = [
        "select#ddlClasificacion",
        "select[id*='Clasificacion']",
        "select[name*='Clasificacion']",
        "select#ddlSubClasificacion",
        "select[id*='SubClasificacion']",
        "input#txtFechaDesde",
        "input[id*='FechaDesde']",
        "button:has-text('Buscar por Estado')",
        "input[type='submit'][value*='Buscar por Estado' i]",
    ]

    started = datetime.now()
    recovered_once = False
    while (datetime.now() - started).total_seconds() * 1000 < RECLAMO_WAIT_MS:
        for target in iter_targets(page):
            try:
                select_count = target.locator("select").count()
                has_search_btn = (
                    target.locator("button:has-text('Buscar por Estado')").count() > 0
                    or target.locator("input[type='submit'][value*='Buscar por Estado' i]").count() > 0
                    or target.locator("input[type='button'][value*='Buscar por Estado' i]").count() > 0
                )
                if select_count >= 2 and has_search_btn:
                    print(
                        "Filtros detectados por estructura "
                        f"(selects={select_count}, buscar={has_search_btn}) en {target.url}"
                    )
                    return
            except PlaywrightError:
                pass

            for selector in selectors:
                try:
                    if target.locator(selector).count() > 0:
                        print(f"Filtros detectados en {target.url} con selector {selector}")
                        return
                except PlaywrightError:
                    continue

        if (not recovered_once) and crmweb_stuck_on_root(page):
            recovered_once = recover_mainframe_reclamo_route(page)
        page.wait_for_timeout(1000)

    dump_debug_artifacts(page, "crm_reclamo_filters_timeout")
    if crmweb_stuck_on_root(page):
        raise RuntimeError(
            "Timeout esperando filtros: CRMWeb quedó en '#/' y no publicó ReclamoCierre. "
            "El token/ruta del iframe no se aplicó correctamente en este entorno."
        )
    raise RuntimeError(
        "Timeout esperando filtros de Reclamo Cierre. "
        "La vista no publicó controles de Clasificacion/SubClasificacion a tiempo."
    )


def open_reclamo_from_menu(page) -> None:
    """Abre Cierre de Reclamos usando menú CRM (Rec.Administrativos -> Cierre de Reclamos)."""
    # Intento 1: interacción visible por hover + click en submenu.
    menu_selectors = [
        "a:has-text('Rec.Administrativos')",
        "#ctl00_MenuBotonera2_oMenun5 a",
        "text=Rec.Administrativos",
    ]
    sub_selectors = [
        "a:has-text('Cierre de Reclamos')",
        "#ctl00_MenuBotonera2_oMenun32 a",
        "a[href*='ReclamoCierre.aspx?SubMenu=475']",
    ]

    for menu_sel in menu_selectors:
        try:
            menu = page.locator(menu_sel).first
            if menu.count() == 0:
                continue
            menu.hover(timeout=UI_TIMEOUT_MS)
            page.wait_for_timeout(400)
            for sub_sel in sub_selectors:
                try:
                    sub = page.locator(sub_sel).first
                    if sub.count() == 0:
                        continue
                    sub.wait_for(state="visible", timeout=2500)
                    sub.click()
                    print(
                        "Navegación a Reclamo Cierre por menú: "
                        f"{menu_sel} -> {sub_sel}"
                    )
                    return
                except PlaywrightError:
                    continue
        except PlaywrightError:
            continue

    # Intento 2: si el submenu no se vuelve visible, toma href y navega directamente.
    target_url = page.evaluate(
        """
        () => {
            const link = Array.from(document.querySelectorAll("a[href*='ReclamoCierre.aspx?SubMenu=475']"))[0];
            return link ? link.href : "";
        }
        """
    )
    if target_url:
        print(f"Navegación a Reclamo Cierre por href de menú: {target_url}")
        page.goto(target_url, wait_until="domcontentloaded")
        return

    raise RuntimeError(
        "No se pudo abrir 'Cierre de Reclamos' desde el menú Rec.Administrativos."
    )


def export_and_download_report(page, download_dir: Path) -> Path:
    """Exporta resultados y descarga el archivo desde el popup de confirmación."""
    context = page.context
    dialog_messages = []

    def _on_dialog(dialog):
        msg = (dialog.message or "").strip()
        dialog_messages.append(msg)
        print(f"Dialog detectado [{dialog.type}]: {msg}")
        try:
            dialog.accept()
        except PlaywrightError:
            pass

    page.on("dialog", _on_dialog)

    export_requests = []

    def _on_request(req):
        try:
            url = (req.url or "").strip()
            if not url:
                return
            low = url.lower()
            if "export" in low or "reporte" in low:
                req_headers = req.headers or {}
                forward_headers = {}
                for hk in [
                    "authorization",
                    "accept",
                    "origin",
                    "referer",
                    "x-requested-with",
                    "x-xsrf-token",
                    "x-csrf-token",
                    "x-auth-token",
                ]:
                    hv = req_headers.get(hk)
                    if hv:
                        forward_headers[hk] = hv
                export_requests.append(
                    {
                        "url": url,
                        "method": (req.method or "GET").upper(),
                        "post_data": req.post_data or "",
                        "content_type": req.header_value("content-type") or "",
                        "headers": forward_headers,
                    }
                )
        except PlaywrightError:
            pass

    page.on("request", _on_request)

    export_selectors = [
        "button:has-text('Exportar Resultados')",
        "a:has-text('Exportar Resultados')",
        "button:has-text('Exportar')",
        "a:has-text('Exportar')",
        "input[type='button'][value*='Exportar' i]",
        "input[type='submit'][value*='Exportar' i]",
        "button[id*='Exportar' i]",
        "a[id*='Exportar' i]",
        "button[class*='export' i]",
        "a[class*='export' i]",
    ]

    def _trigger_export_click() -> bool:
        roots = [p for p in context.pages if not p.is_closed()]
        if click_first_visible_retry(
            page,
            export_selectors,
            "Exportar resultados",
            timeout_ms=12000,
            roots=roots,
        ):
            return True

        # Fallback por JS para variantes de UI con texto/clases no estandar.
        for target in _build_targets_from_roots(roots):
            try:
                clicked = target.evaluate(
                    """
                    () => {
                        const nodes = Array.from(document.querySelectorAll('button,a,input[type="button"],input[type="submit"]'));
                        const match = (el) => {
                            const text = ((el.textContent || '') + ' ' + (el.value || '') + ' ' + (el.id || '') + ' ' + (el.className || '') + ' ' + (el.getAttribute('href') || '')).toLowerCase();
                            return text.includes('exportar') || text.includes('reporte');
                        };
                        const candidate = nodes.find((el) => {
                            if (!match(el)) return false;
                            const style = window.getComputedStyle(el);
                            const visible = style && style.visibility !== 'hidden' && style.display !== 'none';
                            return visible;
                        });
                        if (!candidate) return false;
                        candidate.click();
                        return true;
                    }
                    """
                )
                if clicked:
                    print(f"Click en Exportar resultados por fallback JS en {target.url}")
                    return True
            except PlaywrightError:
                continue
        return False

    # Algunas variantes disparan la descarga directo al hacer click en Exportar.
    export_clicked = False
    try:
        with context.expect_event("download", timeout=15000) as direct_download_info:
            if not _trigger_export_click():
                dump_debug_artifacts(page, "crm_export_button_missing")
                raise RuntimeError("No se encontró un botón visible para: Exportar resultados")
            export_clicked = True
        direct_download = direct_download_info.value
        suggested = (direct_download.suggested_filename or "reporte_cierre_ra.xlsx").strip()
        destination = download_dir / suggested
        direct_download.save_as(str(destination))
        print(f"Reporte descargado directo desde Exportar en: {destination}")
        try:
            page.remove_listener("dialog", _on_dialog)
        except PlaywrightError:
            pass
        return destination
    except PlaywrightTimeoutError:
        if not export_clicked:
            # Puede timeout en caso de click sin descarga; asegura que al menos se haya intentado exportar.
            export_clicked = _trigger_export_click()
            if not export_clicked:
                dump_debug_artifacts(page, "crm_export_button_missing")
                raise RuntimeError("No se encontró un botón visible para: Exportar resultados")
        print("No hubo descarga directa tras Exportar; intentando flujo con popup 'Descargar'.")

    # Flujo esperado por UI: primero "Generando..." y luego link de descarga.
    exact_download_selectors = [
        "a.btn.btn-reporte.btn-block[download][href*='ReporteCierreReclamosAdministrativos.xlsx']",
        "a.btn-reporte.btn-block[download][href*='/reportes/']",
        "a[target='_blank'][download][href*='/reportes/ReporteCierreReclamosAdministrativos.xlsx']",
        "a.btn-reporte:has-text('Descargar')",
    ]

    # Evita bloqueos largos: usamos un timeout acotado para el link visual y
    # priorizamos fallback por request cuando ya existe tráfico de export.
    staged_wait_ms = max(30000, min(DOWNLOAD_TIMEOUT_MS, 90000))
    staged_started = datetime.now()
    staged_clicked = False

    print("Esperando link de descarga de reporte (fase Generando -> Descargar)...")
    while (datetime.now() - staged_started).total_seconds() * 1000 < staged_wait_ms:
        dynamic_roots = [p for p in context.pages if not p.is_closed()]
        dynamic_targets = _build_targets_from_roots(dynamic_roots)

        for target in dynamic_targets:
            for selector in exact_download_selectors:
                try:
                    locator = target.locator(selector)
                    if locator.count() == 0:
                        continue

                    # Puede aparecer antes de quedar clickeable.
                    locator.first.wait_for(state="visible", timeout=1200)

                    href = (locator.first.get_attribute("href") or "").strip()
                    candidate_url = ""
                    if href:
                        if href.startswith("/"):
                            parsed = urlparse(target.url)
                            candidate_url = f"{parsed.scheme}://{parsed.netloc}{href}"
                        else:
                            candidate_url = href

                    try:
                        with context.expect_event("download", timeout=60000) as staged_download_info:
                            locator.first.click()

                        download = staged_download_info.value
                        suggested = (download.suggested_filename or "ReporteCierreReclamosAdministrativos.xlsx").strip()
                        destination = download_dir / suggested
                        download.save_as(str(destination))
                        # Espera corta solicitada para dejar terminar escritura/flush.
                        page.wait_for_timeout(3000)
                        print(f"Reporte descargado desde link 'Descargar' en: {destination}")
                        try:
                            page.remove_listener("dialog", _on_dialog)
                        except PlaywrightError:
                            pass
                        try:
                            page.remove_listener("request", _on_request)
                        except PlaywrightError:
                            pass
                        return destination
                    except PlaywrightTimeoutError:
                        # Fallback: si el click no dispara evento de download, usa el href real del link.
                        if candidate_url:
                            resp = context.request.get(candidate_url, timeout=30000)
                            if resp.ok:
                                ctype = (resp.headers.get("content-type") or "").lower()
                                if any(
                                    k in ctype
                                    for k in ["application/vnd", "excel", "spreadsheet", "octet-stream"]
                                ):
                                    raw = resp.body()
                                    if raw:
                                        destination = download_dir / "ReporteCierreReclamosAdministrativos.xlsx"
                                        destination.write_bytes(raw)
                                        page.wait_for_timeout(3000)
                                        print(f"Reporte descargado por href del link 'Descargar' en: {destination}")
                                        try:
                                            page.remove_listener("dialog", _on_dialog)
                                        except PlaywrightError:
                                            pass
                                        try:
                                            page.remove_listener("request", _on_request)
                                        except PlaywrightError:
                                            pass
                                        return destination
                        continue
                except PlaywrightTimeoutError:
                    continue
                except PlaywrightError:
                    continue

        if not staged_clicked:
            # Mensaje único para evitar ruido en logs.
            staged_clicked = True
            print("Aun no aparece el link 'Descargar'; continuando espera...")

        page.wait_for_timeout(1200)

    def _looks_like_binary_report(content_type: str, payload: bytes) -> bool:
        ctype = (content_type or "").lower()
        if any(k in ctype for k in ["application/vnd", "excel", "spreadsheet", "octet-stream", "csv"]):
            return True
        # Fallback por firma ZIP (xlsx) o CSV/texto con separadores.
        if payload.startswith(b"PK\x03\x04"):
            return True
        head = payload[:512]
        if b";" in head or b"," in head:
            return b"\n" in head
        return False

    def _extract_download_url_from_text(text: str) -> str:
        if not text:
            return ""
        # Busca rutas conocidas del repositorio de reportes o cualquier URL http(s).
        url_match = re.search(r"https?://[^\s\"'<>]+", text)
        if url_match:
            return url_match.group(0)
        path_match = re.search(r"(/reportes?/[^\s\"'<>]+)", text, flags=re.IGNORECASE)
        if path_match:
            return path_match.group(1)
        return ""

    def _extract_download_url_from_json(payload) -> str:
        """Busca de forma recursiva una URL/ruta de reporte dentro de un JSON arbitrario."""
        if isinstance(payload, str):
            return _extract_download_url_from_text(payload)
        if isinstance(payload, dict):
            # Prioriza claves típicas, pero también recorre todo el objeto.
            preferred = [
                "url",
                "downloadUrl",
                "download_url",
                "archivo",
                "file",
                "path",
                "reporte",
                "report",
            ]
            for key in preferred:
                if key in payload:
                    found = _extract_download_url_from_json(payload.get(key))
                    if found:
                        return found
            for value in payload.values():
                found = _extract_download_url_from_json(value)
                if found:
                    return found
            return ""
        if isinstance(payload, list):
            for item in payload:
                found = _extract_download_url_from_json(item)
                if found:
                    return found
        return ""

    def _rewrite_url_for_api(raw_url: str):
        """Reescribe host internos a IP objetivo para APIRequestContext (sin host-resolver-rules)."""
        parsed = urlparse(raw_url)
        host = (parsed.hostname or "").strip().lower()
        if not host:
            return raw_url, ""

        target_ip = ""
        crmrep_host = os.getenv("CRMREP_HOST", "crmrep.telecentro.local").strip().lower()
        crmweb_host = os.getenv("CRMWEB_HOST", "crmweb.telecentro.local").strip().lower()
        apicrm_host = os.getenv("APICRM_HOST", "apicrm.telecentro.local").strip().lower()

        if host == crmrep_host:
            target_ip = os.getenv("CRMREP_TARGET_IP", "").strip()
        elif host == crmweb_host:
            target_ip = os.getenv("CRMWEB_TARGET_IP", "").strip()
        elif host == apicrm_host:
            target_ip = os.getenv("APICRM_TARGET_IP", "").strip()

        if not target_ip:
            return raw_url, ""

        port_part = f":{parsed.port}" if parsed.port else ""
        rewritten = raw_url.replace(f"{parsed.scheme}://{parsed.netloc}", f"{parsed.scheme}://{target_ip}{port_part}", 1)
        return rewritten, parsed.netloc

    def _try_replay_export_request() -> Path | None:
        if not export_requests:
            return None
        req_info = export_requests[-1]
        print(f"Intentando fallback por request de export: {req_info['method']} {req_info['url']}")

        def _persist_body(raw_bytes: bytes, ctype_hint: str, source_label: str) -> Path:
            filename = "ReporteCierreReclamosAdministrativos.xlsx"
            if "csv" in (ctype_hint or ""):
                filename = "ReporteCierreReclamosAdministrativos.csv"
            destination = download_dir / filename
            destination.write_bytes(raw_bytes)
            page.wait_for_timeout(3000)
            print(f"Reporte descargado por {source_label} en: {destination}")
            try:
                page.remove_listener("dialog", _on_dialog)
            except PlaywrightError:
                pass
            try:
                page.remove_listener("request", _on_request)
            except PlaywrightError:
                pass
            return destination

        def _try_fetch_via_page(url: str, method: str, post_data: str, req_headers: dict):
            try:
                js_result = page.evaluate(
                    """
                    async ({url, method, postData, headers}) => {
                        try {
                            const resp = await fetch(url, {
                                method,
                                headers,
                                credentials: 'include',
                                body: (method !== 'GET' && method !== 'HEAD' && postData) ? postData : undefined,
                            });
                            const ctype = resp.headers.get('content-type') || '';
                            const status = resp.status;
                            const ok = resp.ok;
                            const buf = await resp.arrayBuffer();
                            const bytes = new Uint8Array(buf);
                            let binary = '';
                            const chunkSize = 0x8000;
                            for (let i = 0; i < bytes.length; i += chunkSize) {
                                const chunk = bytes.subarray(i, i + chunkSize);
                                binary += String.fromCharCode.apply(null, chunk);
                            }
                            const bodyBase64 = btoa(binary);
                            return { ok, status, ctype, bodyBase64 };
                        } catch (e) {
                            return { ok: false, status: 0, ctype: '', error: String(e) };
                        }
                    }
                    """,
                    {
                        "url": url,
                        "method": method,
                        "postData": post_data or "",
                        "headers": req_headers or {},
                    },
                )
                if not js_result:
                    return None
                status = js_result.get("status", 0)
                ctype = (js_result.get("ctype") or "").lower()
                b64 = js_result.get("bodyBase64") or ""
                print(f"Replay in-page response: status={status} content-type={ctype}")
                if b64:
                    raw = base64.b64decode(b64)
                    if raw and _looks_like_binary_report(ctype, raw):
                        return _persist_body(raw, ctype, "replay in-page fetch")
            except Exception as ex:
                print(f"Replay in-page fallo: {type(ex).__name__}: {ex}")
            return None

        try:
            fetch_kwargs = {"method": req_info["method"], "timeout": 60000}
            if req_info["method"] in {"POST", "PUT", "PATCH", "DELETE"} and req_info["post_data"]:
                fetch_kwargs["data"] = req_info["post_data"]
                if req_info["content_type"]:
                    fetch_kwargs["headers"] = {"content-type": req_info["content_type"]}

            if req_info.get("headers"):
                base_headers = fetch_kwargs.get("headers", {})
                base_headers.update(req_info["headers"])
                fetch_kwargs["headers"] = base_headers

            fetch_url, host_header = _rewrite_url_for_api(req_info["url"])
            if host_header:
                headers = fetch_kwargs.get("headers", {})
                headers.setdefault("host", host_header)
                fetch_kwargs["headers"] = headers

            resp = context.request.fetch(fetch_url, **fetch_kwargs)
            status = resp.status
            ctype = (resp.headers.get("content-type") or "").lower()
            raw = resp.body() or b""
            print(f"Replay export response: status={status} content-type={ctype} url={fetch_url}")

            if resp.ok and raw and _looks_like_binary_report(ctype, raw):
                return _persist_body(raw, ctype, "replay request")

            # Si APIRequestContext responde error/no-binario, intenta desde la propia página.
            in_page_file = _try_fetch_via_page(
                req_info["url"],
                req_info["method"],
                req_info["post_data"],
                req_info.get("headers") or {},
            )
            if in_page_file:
                return in_page_file

            # Algunos backends devuelven JSON con URL temporal del archivo.
            reply_url = ""
            if raw:
                body_text = ""
                try:
                    body_text = raw.decode("utf-8", errors="replace")
                except Exception:
                    body_text = ""

                if "json" in ctype and body_text:
                    try:
                        payload = json.loads(body_text)
                        reply_url = _extract_download_url_from_json(payload)
                    except Exception:
                        pass

                if not reply_url and body_text:
                    reply_url = _extract_download_url_from_text(body_text)

            if reply_url:
                parsed = urlparse(req_info["url"])
                if reply_url.startswith("/"):
                    reply_url = f"{parsed.scheme}://{parsed.netloc}{reply_url}"
                print(f"Fallback replay obtuvo URL de descarga: {reply_url}")

                second_url, second_host = _rewrite_url_for_api(reply_url)
                second_headers = {}
                if second_host:
                    second_headers["host"] = second_host
                second = context.request.get(second_url, timeout=60000, headers=second_headers or None)
                second_ct = (second.headers.get("content-type") or "").lower()
                second_raw = second.body() or b""
                print(f"Descarga 2do paso: status={second.status} content-type={second_ct} url={second_url}")
                if second.ok and second_raw and _looks_like_binary_report(second_ct, second_raw):
                    return _persist_body(second_raw, second_ct, "URL devuelta por export")

                # Segundo paso también desde página si el API context no pudo.
                in_page_second = _try_fetch_via_page(reply_url, "GET", "", req_info.get("headers") or {})
                if in_page_second:
                    return in_page_second
        except Exception as ex:
            print(f"Fallback replay fallo con excepcion: {type(ex).__name__}: {ex}")
        return None

    # Primer intento rápido: si el backend ya recibió exportación, no esperar al botón visual.
    replay_file = _try_replay_export_request()
    if replay_file:
        return replay_file

    # Segundo intento al final del flujo visual acotado.
    replay_file = _try_replay_export_request()
    if replay_file:
        return replay_file

    # Rescate final: intenta URL conocida del reporte con cache-buster.
    try:
        ts = int(datetime.now().timestamp())
        candidate_bases = []
        for target in _build_targets_from_roots([p for p in context.pages if not p.is_closed()]):
            try:
                parsed = urlparse(target.url)
                if parsed.scheme and parsed.netloc:
                    base = f"{parsed.scheme}://{parsed.netloc}"
                    if base not in candidate_bases:
                        candidate_bases.append(base)
            except Exception:
                continue

        for host_env in [
            os.getenv("CRMWEB_HOST", "").strip(),
            os.getenv("CRMREP_HOST", "").strip(),
        ]:
            if host_env:
                base = f"http://{host_env}"
                if base not in candidate_bases:
                    candidate_bases.append(base)

        rescue_paths = [
            f"/reportes/ReporteCierreReclamosAdministrativos.xlsx?nocache={ts}",
            f"/ReporteCierreReclamosAdministrativos.xlsx?nocache={ts}",
        ]

        for base in candidate_bases:
            for rpath in rescue_paths:
                rescue_url = f"{base}{rpath}"
                fetch_url, host_header = _rewrite_url_for_api(rescue_url)
                headers = {"cache-control": "no-cache", "pragma": "no-cache"}
                if host_header:
                    headers["host"] = host_header
                try:
                    resp = context.request.get(fetch_url, timeout=20000, headers=headers)
                except PlaywrightError:
                    continue

                ctype = (resp.headers.get("content-type") or "").lower()
                raw = resp.body() or b""
                if resp.ok and raw and _looks_like_binary_report(ctype, raw):
                    destination = download_dir / "ReporteCierreReclamosAdministrativos.xlsx"
                    destination.write_bytes(raw)
                    page.wait_for_timeout(2000)
                    print(f"Reporte descargado por URL de rescate en: {destination} ({rescue_url})")
                    try:
                        page.remove_listener("dialog", _on_dialog)
                    except PlaywrightError:
                        pass
                    try:
                        page.remove_listener("request", _on_request)
                    except PlaywrightError:
                        pass
                    return destination
    except Exception as ex:
        print(f"Fallback URL de rescate fallo: {type(ex).__name__}: {ex}")
    # Si no apareció ni botón ni archivo directo dentro del timeout total, fallamos con debug.
    if dialog_messages:
        print(f"Dialogs detectados durante export: {dialog_messages}")
    dump_debug_artifacts(page, "crm_export_download_button_missing")
    try:
        page.remove_listener("request", _on_request)
    except PlaywrightError:
        pass
    raise RuntimeError(
        "No se detectó el link 'Descargar' ni quedó disponible la URL directa del reporte dentro del timeout."
    )


def trigger_search_by_estado(page) -> None:
    """Dispara 'Buscar por Estado' con selectores robustos y fallback JS."""
    search_selectors = [
        "button:has-text('Buscar por Estado')",
        "a:has-text('Buscar por Estado')",
        "input[type='submit'][value*='Buscar por Estado' i]",
        "input[type='button'][value*='Buscar por Estado' i]",
        "button[id*='BuscarEstado' i]",
        "input[id*='BuscarEstado' i]",
        "button[id*='Estado' i]",
        "input[id*='Estado' i]",
    ]

    if click_first_visible_retry(
        page,
        search_selectors,
        "Buscar por Estado",
        timeout_ms=15000,
    ):
        return

    # Fallback JS para variantes donde el botón queda renderizado sin selector estable.
    for target in iter_targets(page):
        try:
            clicked = target.evaluate(
                """
                () => {
                    const nodes = Array.from(document.querySelectorAll('button,a,input[type="button"],input[type="submit"]'));
                    const match = (el) => {
                        const text = ((el.textContent || '') + ' ' + (el.value || '') + ' ' + (el.id || '') + ' ' + (el.className || '')).toLowerCase();
                        return text.includes('buscar por estado') || (text.includes('buscar') && text.includes('estado'));
                    };
                    const candidate = nodes.find((el) => {
                        if (!match(el)) return false;
                        const style = window.getComputedStyle(el);
                        return style && style.visibility !== 'hidden' && style.display !== 'none';
                    });
                    if (!candidate) return false;
                    candidate.click();
                    return true;
                }
                """
            )
            if clicked:
                print(f"Click en Buscar por Estado por fallback JS en {target.url}")
                return
        except PlaywrightError:
            continue

    dump_debug_artifacts(page, "crm_buscar_estado_missing")
    raise RuntimeError("No se encontró un botón visible para: Buscar por Estado")

def launch_browser(playwright: Playwright):
    """Prueba varias estrategias de arranque para evitar fallos por deps faltantes."""
    no_sandbox = os.getenv("TELECENTRO_NO_SANDBOX", "0").strip() == "1"
    headless_env = os.getenv("TELECENTRO_HEADLESS", "").strip().lower()
    has_display = bool(os.getenv("DISPLAY", "").strip())

    # Si no hay X server, forzamos headless para poder correr en servidores.
    if headless_env in {"1", "true", "yes"}:
        headless = True
    elif headless_env in {"0", "false", "no"}:
        headless = False
    else:
        headless = not has_display

    launch_kwargs = {"headless": headless, "slow_mo": 120}

    # Forzamos resolución interna de crmweb cuando el DNS local apunta afuera.
    browser_args = []
    host_map_rules = []
    host_map_mode = os.getenv("CRM_FORCE_HOST_MAP", "auto").strip().lower()
    crmweb_host = os.getenv("CRMWEB_HOST", "crmweb.telecentro.local").strip()
    apicrm_host = os.getenv("APICRM_HOST", "apicrm.telecentro.local").strip()
    crmrep_host = os.getenv("CRMREP_HOST", "crmrep.telecentro.local").strip()
    crm_internal_host = os.getenv("CRM_INTERNAL_HOST", "crm.telecentro.local").strip()
    crmweb_target_ip = os.getenv("CRMWEB_TARGET_IP", "").strip()
    apicrm_target_ip = os.getenv("APICRM_TARGET_IP", "").strip()
    crmrep_target_ip = os.getenv("CRMREP_TARGET_IP", "").strip()

    def is_private_ip(ip: str) -> bool:
        try:
            return ipaddress.ip_address(ip).is_private
        except ValueError:
            return False

    if host_map_mode not in {"0", "false", "no", "off"}:
        crmweb_ip = ""
        internal_ip = ""
        need_map = False

        try:
            crmweb_ip = socket.gethostbyname(crmweb_host)
            print(f"Resolución {crmweb_host}: {crmweb_ip}")
        except OSError as err:
            print(f"No se pudo resolver {crmweb_host}: {err}")

        try:
            internal_ip = socket.gethostbyname(crm_internal_host)
            print(f"Resolución {crm_internal_host}: {internal_ip}")
        except OSError as err:
            print(f"No se pudo resolver {crm_internal_host}: {err}")

        target_ip = crmweb_target_ip or internal_ip
        if target_ip and not crmweb_target_ip:
            os.environ["CRMWEB_TARGET_IP"] = target_ip

        if host_map_mode in {"1", "true", "yes", "on"}:
            need_map = bool(target_ip)
        else:
            # Modo auto: aplica map cuando crmweb no resuelve, apunta fuera de red
            # interna o difiere del host interno del CRM.
            if not crmweb_ip and target_ip:
                need_map = True
            elif crmweb_ip and target_ip:
                need_map = (not is_private_ip(crmweb_ip)) or (crmweb_ip != target_ip)

        if need_map and target_ip:
            host_map_rules.append(f"MAP {crmweb_host} {target_ip}")
            print(
                "Aplicando host map CRM: "
                f"{crmweb_host} ({crmweb_ip or 'sin-resolver'}) -> {target_ip}"
            )
        elif host_map_mode in {"1", "true", "yes", "on"} and not target_ip:
            print("CRM_FORCE_HOST_MAP está activo pero no hay IP objetivo válida para mapear.")
        else:
            print("Host map CRM no requerido según resolución actual.")

        # También mapea API CRM (permisos/usuario) cuando el resolver local no la alcanza.
        api_ip = ""
        try:
            api_ip = socket.gethostbyname(apicrm_host)
            print(f"Resolución {apicrm_host}: {api_ip}")
        except OSError as err:
            print(f"No se pudo resolver {apicrm_host}: {err}")

        api_target_ip = apicrm_target_ip or crmweb_target_ip or internal_ip
        if api_target_ip and not apicrm_target_ip:
            os.environ["APICRM_TARGET_IP"] = api_target_ip
        api_need_map = False
        if host_map_mode in {"1", "true", "yes", "on"}:
            api_need_map = bool(api_target_ip)
        else:
            if not api_ip and api_target_ip:
                api_need_map = True
            elif api_ip and api_target_ip:
                api_need_map = (not is_private_ip(api_ip)) or (api_ip != api_target_ip)

        if api_need_map and api_target_ip:
            host_map_rules.append(f"MAP {apicrm_host} {api_target_ip}")
            print(
                "Aplicando host map API CRM: "
                f"{apicrm_host} ({api_ip or 'sin-resolver'}) -> {api_target_ip}"
            )

        # También mapea API de reportes para exportación de cierres.
        rep_ip = ""
        try:
            rep_ip = socket.gethostbyname(crmrep_host)
            print(f"Resolución {crmrep_host}: {rep_ip}")
        except OSError as err:
            print(f"No se pudo resolver {crmrep_host}: {err}")

        rep_target_ip = crmrep_target_ip or crmweb_target_ip or internal_ip
        if rep_target_ip and not crmrep_target_ip:
            os.environ["CRMREP_TARGET_IP"] = rep_target_ip
        rep_need_map = False
        if host_map_mode in {"1", "true", "yes", "on"}:
            rep_need_map = bool(rep_target_ip)
        else:
            if not rep_ip and rep_target_ip:
                rep_need_map = True
            elif rep_ip and rep_target_ip:
                rep_need_map = (not is_private_ip(rep_ip)) or (rep_ip != rep_target_ip)

        if rep_need_map and rep_target_ip:
            host_map_rules.append(f"MAP {crmrep_host} {rep_target_ip}")
            print(
                "Aplicando host map API Reportes: "
                f"{crmrep_host} ({rep_ip or 'sin-resolver'}) -> {rep_target_ip}"
            )

    if host_map_rules:
        host_map_rules.append("EXCLUDE localhost")
        browser_args.append(f"--host-resolver-rules={','.join(host_map_rules)}")

    if no_sandbox:
        browser_args.append("--no-sandbox")

    if browser_args:
        launch_kwargs["args"] = browser_args

    if headless:
        print("Modo headless activo.")

    attempts = []

    # 1) Chromium administrado por Playwright.
    attempts.append(("Playwright Chromium", lambda: playwright.chromium.launch(**launch_kwargs)))

    # 2) Canal Chrome del sistema (si existe).
    attempts.append(("Canal Chrome", lambda: playwright.chromium.launch(channel="chrome", **launch_kwargs)))

    # 3) Ejecutable explícito por variable.
    explicit_exe = os.getenv("TELECENTRO_BROWSER_EXECUTABLE", "").strip()
    if explicit_exe:
        attempts.append(
            (
                f"Ejecutable definido ({explicit_exe})",
                lambda exe=explicit_exe: playwright.chromium.launch(executable_path=exe, **launch_kwargs),
            )
        )

    # 4) Búsqueda automática de binarios comunes.
    for candidate in ["google-chrome", "chromium-browser", "chromium", "/snap/bin/chromium"]:
        exe = shutil.which(candidate) if not candidate.startswith("/") else candidate
        if exe and os.path.exists(exe):
            attempts.append(
                (
                    f"Binario detectado ({exe})",
                    lambda exe=exe: playwright.chromium.launch(executable_path=exe, **launch_kwargs),
                )
            )

    last_error = None
    for label, attempt in attempts:
        try:
            print(f"Intentando abrir navegador con: {label}")
            return attempt()
        except PlaywrightError as err:
            last_error = err
            print(f"Fallo en {label}: {err}")

    print("No se pudo iniciar ningún navegador con las opciones disponibles.")
    print("Si estás en Linux, instala dependencias con: sudo playwright install-deps")
    print("Opcional: export TELECENTRO_BROWSER_EXECUTABLE='/ruta/al/chrome'")
    raise last_error


def run(playwright: Playwright) -> None:
    telecentro_user = get_env("TELECENTRO_USER")
    telecentro_pass = get_env("TELECENTRO_PASS")
    crm_user = get_env("CRM_USER")
    crm_pass = get_env("CRM_PASS")
    crm_internal_host = os.getenv("CRM_INTERNAL_HOST", "crm.telecentro.local").strip() or "crm.telecentro.local"
    crm_url = os.getenv("CRM_URL", CRM_URL_DEFAULT).strip() or CRM_URL_DEFAULT
    crm_reclamo_url = os.getenv("CRM_RECLAMO_URL", CRM_RECLAMO_URL_DEFAULT).strip() or CRM_RECLAMO_URL_DEFAULT
    download_dir = Path(os.getenv("TELECENTRO_DOWNLOAD_DIR", "downloads").strip() or "downloads")
    classifications = get_target_classifications()
    download_dir.mkdir(parents=True, exist_ok=True)

    print(f"Clasificaciones objetivo: {classifications}")

    browser = launch_browser(playwright)
    context = browser.new_context(accept_downloads=True)
    page = context.new_page()

    page.goto(LOGIN_URL, wait_until="domcontentloaded")

    # Completa credenciales con los ids reales del formulario.
    page.fill("#usuario", telecentro_user)
    page.fill("#password", telecentro_pass)

    # Selecciona "Active Directory Corpo" (valor AD-CORPO).
    page.select_option("#validador", value="AD-CORPO")

    # Envía el formulario.
    page.click("#btn-login")

    try:
        page.wait_for_load_state("networkidle", timeout=15000)
        print("Login Telecentro ejecutado.")
    except PlaywrightTimeoutError:
        print("Se hizo click en login de Telecentro, pero la navegación tardó más de lo esperado.")

    # Paso 2: login en CRM.
    try:
        page.goto(crm_url, wait_until="domcontentloaded")
    except PlaywrightError as err:
        print(f"No se pudo abrir CRM URL ({crm_url}): {err}")
        print("Verifica DNS/ruta interna o exporta CRM_URL con una URL accesible desde este host.")
        context.close()
        browser.close()
        sys.exit(1)

    fill_first_visible(
        page,
        [
            "input#UserName",
            "input#txtUsuario",
            "input[id*='UserName']",
            "input[name*='UserName']",
            "input[id*='Usuario']",
            "input[name*='Usuario']",
            "input[placeholder*='Usuario' i]",
            "input[type='text']",
        ],
        crm_user,
        "usuario CRM",
    )
    fill_first_visible(
        page,
        [
            "input#Password",
            "input#txtPassword",
            "input[id*='Password']",
            "input[name*='Password']",
            "input[type='password']",
        ],
        crm_pass,
        "password CRM",
    )
    click_first_visible(
        page,
        [
            "button:has-text('Iniciar sesión')",
            "button:has-text('Iniciar sesion')",
            "input[type='submit'][value*='Iniciar' i]",
            "button[type='submit']",
            "input[type='submit']",
        ],
        "submit CRM",
    )

    try:
        page.wait_for_load_state("networkidle", timeout=15000)
        print("Login CRM ejecutado.")
    except PlaywrightTimeoutError:
        print("Se hizo click en login CRM, pero la navegación tardó más de lo esperado.")

    # Si seguimos en la misma pantalla de login, el alta de sesión fue rechazada.
    if page_requires_login(page):
        crm_login_error = get_page_login_error(page)
        if crm_login_error:
            dump_debug_artifacts(page, "crm_initial_login_failed")
            raise RuntimeError(
                "El login inicial de CRM fue rechazado. "
                f"Detalle CRMWeb: {crm_login_error}"
            )

    # Paso 3: abrir Reclamo Cierre desde el menú del CRM.
    try:
        open_reclamo_from_menu(page)
    except RuntimeError as err:
        # Fallback final opcional a URL directa por si la estructura de menú cambia.
        print(f"Fallo apertura por menú: {err}")
        print(f"Reintentando con URL directa de respaldo: {crm_reclamo_url}")
        try:
            page.goto(crm_reclamo_url, wait_until="domcontentloaded")
        except PlaywrightError as nav_err:
            print(f"No se pudo abrir Reclamo URL ({crm_reclamo_url}): {nav_err}")
            context.close()
            browser.close()
            sys.exit(1)

    # Continúa directo con los filtros en Reclamo Cierre.
    try:
        page.wait_for_load_state("networkidle", timeout=UI_TIMEOUT_MS)
    except PlaywrightTimeoutError:
        pass
    page.wait_for_timeout(2000)

    if has_parking_redirect(page):
        dump_debug_artifacts(page, "crm_iframe_parking_redirect")
        raise RuntimeError(
            "El frame interno de Reclamo Cierre cayó en instantfwding.com. "
            "Se necesita resolver crmweb.telecentro.local a un destino interno válido."
        )

    login_iframe_if_needed(page, crm_user, crm_pass)

    if iframe_requires_login(page):
        iframe_error = get_iframe_login_error(page)
        retried = False
        if "index was out of range" in (iframe_error or "").lower():
            retried = try_reclamo_fallback_via_token_url(page, crm_internal_host)
            if retried:
                login_iframe_if_needed(page, crm_user, crm_pass)

        if iframe_requires_login(page):
            dump_debug_artifacts(page, "crm_iframe_not_authenticated")
            detail = f" Detalle CRMWeb: {iframe_error}" if iframe_error else ""
            raise RuntimeError(
                "No se pudo completar el login embebido de mainFrame. "
                "Sigue apareciendo la pantalla de login y no se llega a Clasificacion."
                f"{detail}"
            )

    if page_requires_login(page):
        ok_direct_login = login_page_if_needed(page, crm_user, crm_pass)
        if not ok_direct_login:
            page_error = get_page_login_error(page)
            dump_debug_artifacts(page, "crm_page_login_not_authenticated")
            detail = f" Detalle CRMWeb: {page_error}" if page_error else ""
            raise RuntimeError(
                "La pantalla principal quedó en login de CRMWeb y no se pudo autenticar."
                f"{detail}"
            )

        # Tras login directo, vuelve a la URL objetivo de reclamo.
        page.goto(crm_reclamo_url, wait_until="domcontentloaded")
        try:
            page.wait_for_load_state("networkidle", timeout=UI_TIMEOUT_MS)
        except PlaywrightTimeoutError:
            pass
        login_iframe_if_needed(page, crm_user, crm_pass)

    if iframe_requires_login(page):
        iframe_error = get_iframe_login_error(page)
        if "index was out of range" in (iframe_error or "").lower():
            dump_debug_artifacts(page, "crmweb_wrong_target_ip")
            raise RuntimeError(
                "CRMWeb responde login legacy con error 'Index was out of range'. "
                "Esto suele indicar que CRMWEB_TARGET_IP no apunta al backend correcto de crmweb. "
                "Configura CRMWEB_TARGET_IP con la IP interna real de crmweb.telecentro.local."
            )

    if crmweb_stuck_on_root(page):
        recover_mainframe_reclamo_route(page)

    wait_for_reclamo_filters(page)

    lookback_days = get_lookback_days(7)
    fecha_hasta = datetime.now().strftime("%d/%m/%Y")
    fecha_desde = (datetime.now() - timedelta(days=lookback_days)).strftime("%d/%m/%Y")
    print(f"Rango de fechas configurado: ultimos {lookback_days} dias ({fecha_desde} -> {fecha_hasta})")

    for idx, classification in enumerate(classifications, start=1):
        print(f"Procesando clasificación {idx}/{len(classifications)}: {classification}")
        try:
            select_first_visible(
                page,
                [
                    "select#ddlClasificacion",
                    "select[id*='Clasificacion']",
                    "select[name*='Clasificacion']",
                ],
                classification,
                "Clasificacion",
            )
        except RuntimeError as err:
            try:
                select_option_in_any_select(page, classification, "Clasificacion")
            except RuntimeError:
                dump_debug_artifacts(page, "crm_clasificacion_error")
                raise RuntimeError(f"No se pudo seleccionar Clasificacion='{classification}'") from err

        # Puede tardar en refrescar subclasificación según el onchange del CRM.
        page.wait_for_timeout(1000)
        try:
            select_first_visible(
                page,
                [
                    "select#ddlSubClasificacion",
                    "select[id*='SubClasificacion']",
                    "select[name*='SubClasificacion']",
                ],
                "-TODOS-",
                "SubClasificacion",
            )
        except RuntimeError:
            select_option_in_any_select(page, "-TODOS-", "SubClasificacion")

        # Nuevo requisito: filtrar por Estado = Cerrado (forzado y validado).
        select_estado_cerrado(page)

        fill_first_visible(
            page,
            [
                "input#txtFechaDesde",
                "input#txtFecDesde",
                "input[id*='FechaDesde']",
                "input[id*='FecDesde']",
                "input[name*='FechaDesde']",
                "input[name*='FecDesde']",
                "input[placeholder*='Desde' i]",
                "xpath=//*[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'fecha desde')]/following::input[1]",
            ],
            fecha_desde,
            "Fecha Desde",
        )
        try:
            fill_first_visible(
                page,
                [
                    "input#txtFechaHasta",
                    "input#txtFecHasta",
                    "input[id*='FechaHasta']",
                    "input[id*='FecHasta']",
                    "input[name*='FechaHasta']",
                    "input[name*='FecHasta']",
                    "input[placeholder*='Hasta' i]",
                    "xpath=//*[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'fecha hasta')]/following::input[1]",
                ],
                fecha_hasta,
                "Fecha Hasta",
            )
        except RuntimeError:
            fill_date_input_by_order(page, 1, fecha_hasta, "Fecha Hasta")

        # Refuerzo: en algunas variantes la UI pisa Fecha Hasta con el valor de Desde.
        # Reescribimos ambos campos por orden para dejar el par consistente.
        try:
            fill_date_input_by_order(page, 0, fecha_desde, "Fecha Desde (refuerzo por orden)")
            fill_date_input_by_order(page, 1, fecha_hasta, "Fecha Hasta (refuerzo por orden)")
        except RuntimeError:
            # Si no hay 2 campos detectables por orden, se continúa con la validación normal.
            pass

        ensure_filters_consistency(page, classification, fecha_desde, fecha_hasta)

        trigger_search_by_estado(page)
        wait_results_refresh_after_search(page)
        print(
            f"Búsqueda ejecutada para '{classification}' con rango "
            f"{fecha_desde} -> {fecha_hasta}."
        )

        downloaded = export_and_download_report(page, download_dir)
        suffix = downloaded.suffix or ".xlsx"
        final_name = f"{classification_filename(classification)}{suffix}"
        renamed = downloaded.with_name(final_name)
        if renamed.exists():
            renamed.unlink()
        downloaded.rename(renamed)
        print(f"Reporte final para clasificación '{classification}': {renamed}")

        extract_original_csv_from_report(renamed, classification, download_dir)
        extract_filtered_csv_from_report(renamed, classification, download_dir)

        # El entregable final queda en CSV; el XLSX intermedio se elimina.
        if renamed.exists():
            renamed.unlink()

    # Si queres inspeccionar manualmente, habilita TELECENTRO_PAUSE=1 con DISPLAY disponible.
    should_pause = os.getenv("TELECENTRO_PAUSE", "0").strip() == "1"
    if should_pause and os.getenv("DISPLAY", "").strip():
        page.pause()
    else:
        context.close()
        browser.close()


if __name__ == "__main__":
    with sync_playwright() as p:
        run(p)
